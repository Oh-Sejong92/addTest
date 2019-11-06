from configparser import ConfigParser
from collections import defaultdict
from openpyxl.styles import Color, PatternFill
from openpyxl.reader.excel import load_workbook
import pandas as pd
import math
import os
import win32api
win32api.MessageBox(0, '프로그램을 실행합니다. \n"'"확인"'" 버튼 을 누른 후 잠시만 기다려 주세요.', '실행 중')


# todo ini파일 읽어 오기
def read_ini(filename, sections, tag):
    """
    ini파일 읽어 오기
    :param filename: 파일명
    :param sections: ini 파일 내 section 명
    :param tag: section 내 tag 명
    :return: section, tag 명
    """
    parser = ConfigParser()
    parser.read(filename)
    return parser.get(sections, tag)


# todo 엑셀 파일 불러오기
def dataset_excelload(target_file_path, sheetname, quarter):
    #  엑셀 파일 유무에 따른 분기
    try:
        df = pd.read_excel(target_file_path, sheet_name=sheetname)
        if quarter == 0:
            df['PLAN_YRMON'] = df['PLAN_YRMON'].astype(str).apply(lambda x: x[:4])
            df['PLAN_YRMON'] = df['PLAN_YRMON'].astype('int64')
            df = df.sort_values(by=['PLAN_YRMON', 'EQP_NIN', 'LOINO', 'TVNO', 'PRCSS_DATE'], axis=0)
            df = df.reset_index(drop=True)
        df_to_dic = df.to_dict('index')
        # pprint.pprint(df_to_dic)
        return df_to_dic
    except Exception:
        print("[ERROR] Excel 파일 로드 실패 : %s" % (target_file_path))
        return None


# todo 데이터셋 딕셔너리로 재정렬 메소드
def rearrangement_dataset(dataset):
    # dataset 재정렬한 딕셔너리
    processing_data = {}
    processing_data2 = {}
    processing_data3 = {}
    final_processing_data = {}
    new_insert_niin_data = {}
    count = 0
    count_dic = {}
    for sub_key in dataset.keys():
        new_insert_niin_data[dataset[sub_key]['EQP_NIN']] = {}
    for sub_key in dataset.keys():
        new_insert_niin_data[dataset[sub_key]['EQP_NIN']][dataset[sub_key]['PLAN_YRMON']] = {}
    for sub_key in dataset.keys():
        new_insert_niin_data[dataset[sub_key]['EQP_NIN']][dataset[sub_key]['PLAN_YRMON']][dataset[sub_key]['LOINO']] = {}
    # 동일한 TVNO에 대한 NIIN을 201에 맞추기 위한 작업
    for sub_key in dataset.keys():
        new_insert_niin_data[dataset[sub_key]['EQP_NIN']][dataset[sub_key]['PLAN_YRMON']][dataset[sub_key]['LOINO']][dataset[sub_key]['TVNO']] = {}
    for sub_key in dataset.keys():
        new_insert_niin_data[dataset[sub_key]['EQP_NIN']][dataset[sub_key]['PLAN_YRMON']][
            dataset[sub_key]['LOINO']][dataset[sub_key]['TVNO']][dataset[sub_key]['TIC']] = {}
    for sub_key, sub_value in dataset.items():
        if not new_insert_niin_data[dataset[sub_key]['EQP_NIN']][dataset[sub_key]['PLAN_YRMON']][dataset[sub_key]['LOINO']][dataset[sub_key]['TVNO']][dataset[sub_key]['TIC']]:
            new_insert_niin_data[dataset[sub_key]['EQP_NIN']][dataset[sub_key]['PLAN_YRMON']][dataset[sub_key]['LOINO']][dataset[sub_key]['TVNO']][dataset[sub_key]['TIC']][1] = sub_value
        else:
            i = len(new_insert_niin_data[dataset[sub_key]['EQP_NIN']][dataset[sub_key]['PLAN_YRMON']][
                dataset[sub_key]['LOINO']][dataset[sub_key]['TVNO']][dataset[sub_key]['TIC']])
            new_insert_niin_data[dataset[sub_key]['EQP_NIN']][dataset[sub_key]['PLAN_YRMON']][dataset[sub_key]['LOINO']][dataset[sub_key]['TVNO']][dataset[sub_key]['TIC']][i+1] = sub_value
    index = 0
    for eqp_nin in new_insert_niin_data.keys():
        for year in new_insert_niin_data[eqp_nin].keys():
            for loino in new_insert_niin_data[eqp_nin][year].keys():
                for tvno in new_insert_niin_data[eqp_nin][year][loino].keys():
                    for tic in new_insert_niin_data[eqp_nin][year][loino][tvno].keys():
                        if tic == 201:
                            n_niin = new_insert_niin_data[eqp_nin][year][loino][tvno][tic][1]['NIIN']
                    for tic2 in new_insert_niin_data[eqp_nin][year][loino][tvno].keys():
                        for quarty_req in new_insert_niin_data[eqp_nin][year][loino][tvno][tic2].keys():
                            new_insert_niin_data[eqp_nin][year][loino][tvno][tic2][quarty_req]['N_NIIN'] = n_niin
                            new_insert_niin_data[eqp_nin][year][loino][tvno][tic2][quarty_req]['QUARTY_REQ'] = quarty_req
                            processing_data[index] = new_insert_niin_data[eqp_nin][year][loino][tvno][tic2][quarty_req]
                            index += 1

    for sub_key in processing_data.keys():
        processing_data2[processing_data[sub_key]['EQP_NIN']] = {}
    for sub_key in processing_data.keys():
        processing_data2[processing_data[sub_key]['EQP_NIN']][processing_data[sub_key]['PLAN_YRMON']] = {}
    for sub_key in processing_data.keys():
        processing_data2[processing_data[sub_key]['EQP_NIN']][processing_data[sub_key]['PLAN_YRMON']][processing_data[sub_key]['LOINO']] = {}
    for sub_key in processing_data.keys():
        processing_data2[processing_data[sub_key]['EQP_NIN']][processing_data[sub_key]['PLAN_YRMON']][processing_data[sub_key]['LOINO']][processing_data[sub_key]['N_NIIN']] = {}
    for sub_key in processing_data.keys():
        processing_data2[processing_data[sub_key]['EQP_NIN']][processing_data[sub_key]['PLAN_YRMON']][processing_data[sub_key]['LOINO']][processing_data[sub_key]['N_NIIN']][processing_data[sub_key]['TVNO']] = {}
    for sub_key in processing_data.keys():
        processing_data2[processing_data[sub_key]['EQP_NIN']][processing_data[sub_key]['PLAN_YRMON']][processing_data[sub_key]['LOINO']][processing_data[sub_key]['N_NIIN']][processing_data[sub_key]['TVNO']][processing_data[sub_key]['TIC']] = {}
    for sub_key, sub_value in processing_data.items():
        processing_data2[processing_data[sub_key]['EQP_NIN']][processing_data[sub_key]['PLAN_YRMON']][processing_data[sub_key]['LOINO']][processing_data[sub_key]['N_NIIN']][processing_data[sub_key]['TVNO']][processing_data[sub_key]['TIC']][processing_data[sub_key]['QUARTY_REQ']] = sub_value
    index = 0
    for eqp_nin in processing_data2.keys():
        for year in processing_data2[eqp_nin].keys():
            for loino in processing_data2[eqp_nin][year].keys():
                for n_niin in processing_data2[eqp_nin][year][loino].keys():
                    i = 1
                    for tvno in processing_data2[eqp_nin][year][loino][n_niin].keys():
                        for tic in processing_data2[eqp_nin][year][loino][n_niin][tvno].keys():
                            for quarty_req in processing_data2[eqp_nin][year][loino][n_niin][tvno][tic].keys():
                                processing_data2[eqp_nin][year][loino][n_niin][tvno][tic][quarty_req]['SEQ'] = i
                                processing_data3[index] = processing_data2[eqp_nin][year][loino][n_niin][tvno][tic][quarty_req]
                                index += 1
                        i += 1

    for sub_key in processing_data3.keys():
        final_processing_data[processing_data3[sub_key]['EQP_NIN']] = {}
    for sub_key in processing_data3.keys():
        final_processing_data[processing_data3[sub_key]['EQP_NIN']][processing_data3[sub_key]['PLAN_YRMON']] = {}
    for sub_key in processing_data3.keys():
        final_processing_data[processing_data3[sub_key]['EQP_NIN']][processing_data3[sub_key]['PLAN_YRMON']][processing_data3[sub_key]['LOINO']] = {}
    for sub_key in processing_data3.keys():
        final_processing_data[processing_data3[sub_key]['EQP_NIN']][processing_data3[sub_key]['PLAN_YRMON']][processing_data3[sub_key]['LOINO']][processing_data3[sub_key]['N_NIIN']] = {}
    for sub_key in processing_data3.keys():
        final_processing_data[processing_data3[sub_key]['EQP_NIN']][processing_data3[sub_key]['PLAN_YRMON']][processing_data3[sub_key]['LOINO']][processing_data3[sub_key]['N_NIIN']][processing_data3[sub_key]['TVNO']] = {}
    for sub_key in processing_data3.keys():
        final_processing_data[processing_data3[sub_key]['EQP_NIN']][processing_data3[sub_key]['PLAN_YRMON']][processing_data3[sub_key]['LOINO']][processing_data3[sub_key]['N_NIIN']][processing_data3[sub_key]['TVNO']][processing_data3[sub_key]['TIC']] = {}
    for sub_key, sub_value in processing_data3.items():
        final_processing_data[processing_data3[sub_key]['EQP_NIN']][processing_data3[sub_key]['PLAN_YRMON']][processing_data3[sub_key]['LOINO']][processing_data3[sub_key]['N_NIIN']][processing_data3[sub_key]['TVNO']][processing_data3[sub_key]['TIC']][processing_data3[sub_key]['QUARTY_REQ']] = sub_value

    # 연도별 거래증비서 수 & 총 수
    for eqp_nin in final_processing_data.keys():
        for year in final_processing_data[eqp_nin].keys():
            count_dic[year] = []
            for loino in final_processing_data[eqp_nin][year].keys():
                for n_niin in final_processing_data[eqp_nin][year][loino].keys():
                    for tvno in final_processing_data[eqp_nin][year][loino][n_niin].keys():
                        count_dic[year].append(tvno)
                        count = count + 1
            count_dic[year] = len(count_dic[year])
    return final_processing_data, count_dic, count


# todo 연도별 국가공휴일 정렬 메소드
def holiday_processing(holiday_dataset):
    holiday_processing_data = {}
    for sub_key in holiday_dataset.keys():
        holiday_processing_data[holiday_dataset[sub_key]['연도']] = []
    for sub_key in holiday_dataset.keys():
        holiday_processing_data[holiday_dataset[sub_key]['연도']].append(str(holiday_dataset[sub_key]['날짜']))
    return holiday_processing_data


# todo 공휴일 제외한 날짜 계산 메소드
def working_dates(start_date, end_date, holiday_processing_dataset):
    dt_index = pd.date_range(start=start_date, end=end_date, freq='B')  # 시작일자에서 종료일자까지의 주말을 제외한 평일(pandas 활용)
    repair_term = dt_index.size  # 기간 내 평일일자의 수
    # 기간내 연도 변동이 있었는가
    if int(start_date[0:4]) == int(end_date[0:4]):
        for i in dt_index:
            if i in holiday_processing_dataset[int(start_date[0:4])]:
                repair_term -= 1
    else:
        for i in dt_index:
            if i in holiday_processing_dataset[int(start_date[0:4])]:
                repair_term -= 1
        for i in dt_index:
            if i in holiday_processing_dataset[int(start_date[0:4])]:
                repair_term -= 1
    return repair_term


# todo 총 정비기간 분석 메소드
def compare_temp(processing_dataset, holiday_processing_dataset):
    d_index = 0
    for eqp_nin in processing_dataset.keys():
        for year in processing_dataset[eqp_nin].keys():
            for loino in processing_dataset[eqp_nin][year].keys():
                for niin in processing_dataset[eqp_nin][year][loino].keys():
                    for tvno_key in processing_dataset[eqp_nin][year][loino][niin].keys():
                        c_qty = 0
                        if 132 in processing_dataset[eqp_nin][year][loino][niin][tvno_key]:
                            for quarty_req in processing_dataset[eqp_nin][year][loino][niin][tvno_key][132].keys():
                                c_qty += processing_dataset[eqp_nin][year][loino][niin][tvno_key][132][quarty_req][
                                    'QTY']
                        if 133 in processing_dataset[eqp_nin][year][loino][niin][tvno_key]:
                            for quarty_req in processing_dataset[eqp_nin][year][loino][niin][tvno_key][133].keys():
                                c_qty += processing_dataset[eqp_nin][year][loino][niin][tvno_key][133][quarty_req][
                                    'QTY']
                        for tic in processing_dataset[eqp_nin][year][loino][niin][tvno_key].keys():
                            for quarty_req, values in processing_dataset[eqp_nin][year][loino][niin][tvno_key][tic].items():
                                processing_dataset[eqp_nin][year][loino][niin][tvno_key][tic][quarty_req]['50%_SAT'] = math.ceil(c_qty*50/100)
                                processing_dataset[eqp_nin][year][loino][niin][tvno_key][tic][quarty_req]['75%_SAT'] = math.ceil(c_qty*75/100)
                                processing_dataset[eqp_nin][year][loino][niin][tvno_key][tic][quarty_req]['95%_SAT'] = math.ceil(c_qty*95/100)
                        r_cwt = 0
                        r_qty = 0
                        r_cwt50 = 0
                        r_cwt75 = 0
                        r_cwt95 = 0
                        tvno = processing_dataset[eqp_nin][year][loino][niin][tvno_key][201][1]['TVNO']
                        start_date = str(processing_dataset[eqp_nin][year][loino][niin][tvno_key][201][1]['PRCSS_DATE']) # 시작 날짜
                        processing_dataset[eqp_nin][year][loino][niin][tvno_key][201][1]['TOTAL_TERM'] = 0
                        bol50 = True
                        bol75 = True
                        bol95 = True
                        if 132 in processing_dataset[eqp_nin][year][loino][niin][tvno_key]:
                            if 232 in processing_dataset[eqp_nin][year][loino][niin][tvno_key]:
                                for quarty_req in processing_dataset[eqp_nin][year][loino][niin][tvno_key][232].keys():
                                    processing_dataset[eqp_nin][year][loino][niin][tvno_key][232][quarty_req]['TOTAL_TERM'] = 0
                            for quarty_req, values in processing_dataset[eqp_nin][year][loino][niin][tvno_key][132].items():
                                end_date = str(processing_dataset[eqp_nin][year][loino][niin][tvno_key][132][quarty_req]['PRCSS_DATE'])  # TIC - 132 // 종료 날짜
                                repair_term = working_dates(start_date, end_date, holiday_processing_dataset)
                                processing_dataset[eqp_nin][year][loino][niin][tvno_key][132][quarty_req]['TOTAL_TERM'] = repair_term
                                qty = processing_dataset[eqp_nin][year][loino][niin][tvno_key][132][quarty_req]['QTY']
                                r_qty += qty
                                r_cwt += repair_term*qty
                                if bol50:
                                    if r_qty >= processing_dataset[eqp_nin][year][loino][niin][tvno_key][132][quarty_req]['50%_SAT']:
                                        r_cwt50 = r_cwt / r_qty
                                        bol50 = False
                                if bol75:
                                    if r_qty >= processing_dataset[eqp_nin][year][loino][niin][tvno_key][132][quarty_req]['75%_SAT']:
                                        r_cwt75 = r_cwt / r_qty
                                        bol75 = False
                                if bol95:
                                    if r_qty >= processing_dataset[eqp_nin][year][loino][niin][tvno_key][132][quarty_req]['95%_SAT']:
                                        r_cwt95 = r_cwt / r_qty
                                        bol95 = False
                        if 133 in processing_dataset[eqp_nin][year][loino][niin][tvno_key]:
                            if 221 in processing_dataset[eqp_nin][year][loino][niin][tvno_key]:
                                for quarty_req in processing_dataset[eqp_nin][year][loino][niin][tvno_key][221].keys():
                                    processing_dataset[eqp_nin][year][loino][niin][tvno_key][221][quarty_req]['TOTAL_TERM'] = 0
                            if 233 in processing_dataset[eqp_nin][year][loino][niin][tvno_key]:
                                for quarty_req in processing_dataset[eqp_nin][year][loino][niin][tvno_key][233].keys():
                                    processing_dataset[eqp_nin][year][loino][niin][tvno_key][233][quarty_req]['TOTAL_TERM'] = 0
                            for quarty_req, values in processing_dataset[eqp_nin][year][loino][niin][tvno_key][133].items():
                                end_date = str(processing_dataset[eqp_nin][year][loino][niin][tvno_key][133][quarty_req]['PRCSS_DATE'])  # TIC - 133 // 종료 날짜
                                repair_term = working_dates(start_date, end_date, holiday_processing_dataset)
                                processing_dataset[eqp_nin][year][loino][niin][tvno_key][133][quarty_req]['TOTAL_TERM'] = repair_term
                                qty = processing_dataset[eqp_nin][year][loino][niin][tvno_key][133][quarty_req]['QTY']
                                r_cwt += repair_term*qty
                                r_qty += qty
                                if bol50:
                                    if r_qty >= processing_dataset[eqp_nin][year][loino][niin][tvno_key][133][quarty_req]['50%_SAT']:
                                        r_cwt50 = r_cwt / r_qty
                                        bol50 = False
                                if bol75:
                                    if r_qty >= processing_dataset[eqp_nin][year][loino][niin][tvno_key][133][quarty_req]['75%_SAT']:
                                        r_cwt75 = r_cwt / r_qty
                                        bol75 = False
                                if bol95:
                                    if r_qty >= processing_dataset[eqp_nin][year][loino][niin][tvno_key][133][quarty_req]['95%_SAT']:
                                        r_cwt95 = r_cwt / r_qty
                                        bol95 = False
                        result_cwt = r_cwt/r_qty
                        for tic in processing_dataset[eqp_nin][year][loino][niin][tvno_key].keys():
                            for quarty_req in processing_dataset[eqp_nin][year][loino][niin][tvno_key][tic].keys():
                                processing_dataset[eqp_nin][year][loino][niin][tvno_key][tic][quarty_req]['CWT'] = result_cwt
                                processing_dataset[eqp_nin][year][loino][niin][tvno_key][tic][quarty_req]['50%_CWT'] = r_cwt50
                                processing_dataset[eqp_nin][year][loino][niin][tvno_key][tic][quarty_req]['75%_CWT'] = r_cwt75
                                processing_dataset[eqp_nin][year][loino][niin][tvno_key][tic][quarty_req]['95%_CWT'] = r_cwt95
    final_processing_dataset = {}
    for eqp_nin in processing_dataset.keys():
        for year in processing_dataset[eqp_nin].keys():
            for loino in processing_dataset[eqp_nin][year].keys():
                for niin in processing_dataset[eqp_nin][year][loino].keys():
                    for tvno_key in processing_dataset[eqp_nin][year][loino][niin].keys():
                        for tic in processing_dataset[eqp_nin][year][loino][niin][tvno_key].keys():
                            for values in processing_dataset[eqp_nin][year][loino][niin][tvno_key][tic].values():
                                final_processing_dataset[d_index] = values
                                d_index += 1
    processing_df = pd.DataFrame(final_processing_dataset)
    processing_df = processing_df.T
    processing_df['REQ_UC'] = processing_df['REQ_UC'].astype('int64')
    processing_df['PRCSS_DATE'] = processing_df['PRCSS_DATE'].astype('int64')
    processing_df['EQP_NIN'] = processing_df['EQP_NIN'].astype('int64')
    processing_df['QTY'] = processing_df['QTY'].astype('int64')
    processing_df['SEQ'] = processing_df['SEQ'].astype('int64')
    processing_df['TIC'] = processing_df['TIC'].astype('int64')
    processing_df['QUARTY_REQ'] = processing_df['QUARTY_REQ'].astype('int64')
    processing_df['TOTAL_TERM'] = processing_df['TOTAL_TERM'].astype('int64')
    processing_df['50%_SAT'] = processing_df['50%_SAT'].astype('int64')
    processing_df['75%_SAT'] = processing_df['75%_SAT'].astype('int64')
    processing_df['95%_SAT'] = processing_df['95%_SAT'].astype('int64')
    processing_df['50%_CWT'] = processing_df['50%_CWT'].astype('int64')
    processing_df['75%_CWT'] = processing_df['75%_CWT'].astype('int64')
    processing_df['95%_CWT'] = processing_df['95%_CWT'].astype('int64')
    processing_df['CWT'] = processing_df['CWT'].astype('int64')
    return processing_df


# todo 연도별 청구부품 수
def yearly_niin_count(processing_df):
    l = []
    d = defaultdict(list)
    for n_niin, group in processing_df.groupby([processing_df['PLAN_YRMON'], processing_df['LOINO'], processing_df["N_NIIN"]]):
        l.append(n_niin)
    for v, y, k in l:
        d[v].append(k)
    yearly_niin_count = {}
    for key, value in d.items():
        yearly_niin_count[key] = len(value)
    return yearly_niin_count


# todo 요청사항 테이블 만들기
def finally_dataframe(processing_df, yearly_niin_count, yearly_certificate_count):
    tvno_grouped_df = processing_df.drop_duplicates(['EQP_NIN', 'PLAN_YRMON',  'LOINO', 'N_NIIN', 'TVNO'], keep="last")
    yearly_cwt_grouped = tvno_grouped_df['CWT'].groupby(tvno_grouped_df['PLAN_YRMON'])
    yearly_50persent_cwt_grouped = tvno_grouped_df['50%_CWT'].groupby(tvno_grouped_df['PLAN_YRMON'])
    yearly_75persent_cwt_grouped = tvno_grouped_df['75%_CWT'].groupby(tvno_grouped_df['PLAN_YRMON'])
    yearly_95persent_cwt_grouped = tvno_grouped_df['95%_CWT'].groupby(tvno_grouped_df['PLAN_YRMON'])
    yearly_avg = round(yearly_cwt_grouped.mean(), 2)  # CWT 평균
    yearly_std = round(yearly_cwt_grouped.std(), 2)  # CWT 표준편차
    yearly_niin_scwt_df = pd.DataFrame({'PLAN_YRMON': [], '50%_CWT': [], '75%_CWT': [], '95%_CWT': []})
    key = []
    for (key1, key2, key3, key4), group in tvno_grouped_df.groupby(['EQP_NIN', 'PLAN_YRMON',  'LOINO', 'N_NIIN']):
        group = group.groupby('N_NIIN').mean()
        yearly_niin_scwt_df = pd.concat((yearly_niin_scwt_df, group), join='inner', ignore_index=True)
        key.append(key2)
    yearly_niin_scwt_df['PLAN_YRMON'] = key
    yearly_niin_50p_cwt_grouped = yearly_niin_scwt_df['50%_CWT'].groupby(yearly_niin_scwt_df['PLAN_YRMON'])
    yearly_niin_75p_cwt_grouped = yearly_niin_scwt_df['75%_CWT'].groupby(yearly_niin_scwt_df['PLAN_YRMON'])
    yearly_niin_95p_cwt_grouped = yearly_niin_scwt_df['95%_CWT'].groupby(yearly_niin_scwt_df['PLAN_YRMON'])
    niin_quantile50 = round(yearly_niin_50p_cwt_grouped.mean(), 4)  # "50% 만족 CWT/누적청구"
    niin_quantile75 = round(yearly_niin_75p_cwt_grouped.mean(), 4)  # "75% 만족 CWT/누적청구"
    niin_quantile95 = round(yearly_niin_95p_cwt_grouped.mean(), 4)  # "95% 만족 CWT/누적청구"

    finally_df = pd.DataFrame(columns=[2013, 2014, 2015, 2016, 2017, 2018], index=['총 청구 부품 수', '총 청구 건수', 'CWT 평균(일)', '표준편차(일)', '50% 만족 CWT / 누적청구건수(일)', '75% 만족 CWT / 누적청구건수(일)', '95% 만족 CWT / 누적청구건수(일)'])
    finally_df.loc['총 청구 부품 수'] = yearly_niin_count
    finally_df.loc['총 청구 건수'] = yearly_certificate_count
    finally_df.loc['CWT 평균(일)'] = yearly_avg
    finally_df.loc['표준편차(일)'] = yearly_std
    finally_df.loc['50% 만족 CWT / 누적청구건수(일)'] = niin_quantile50
    finally_df.loc['75% 만족 CWT / 누적청구건수(일)'] = niin_quantile75
    finally_df.loc['95% 만족 CWT / 누적청구건수(일)'] = niin_quantile95
    return finally_df


# todo 파일생성 경로가 없을시 하위 디렉토리 생성
def mkdir_if_not(filepath):
    """
    :param filepath: 파일경로
    :return: 없음
    """
    if not os.path.exists(filepath):
        os.makedirs(filepath)


def count_grouping_yearly(yearly_n_niin_grouped_df, when_complet_days, p_str, i):
    if i == 0:  # 총 계
        years = when_complet_days.unstack().unstack().index
        for year in years:
            yearly_n_niin_grouped_df.loc[p_str][year] = len(when_complet_days[year])
    else:   # 각 일 수별 계산값
        years = when_complet_days.unstack().unstack().index
        for year in years:
            yearly_n_niin_grouped_df.loc[p_str][year] = str(len(when_complet_days[year]))


def excel_write_fd(finally_data):
    f_writer = pd.ExcelWriter('./ResultFile/finally_table.xlsx', engine='openpyxl')
    finally_data.to_excel(f_writer, sheet_name='Sheet1')
    f_writer.save()


# todo 요청사항 테이블 엑셀 저장
def excel_write_finally_df(finally_data):
    e_writer = pd.ExcelWriter('./ResultFile/response_table.xlsx', engine='openpyxl')
    finally_data.to_excel(e_writer, sheet_name='Sheet1')
    e_writer.save()

    book1 = load_workbook('./ResultFile/response_table.xlsx')
    ws1 = book1.worksheets[0]
    ws1.column_dimensions['A'].width = 33
    book1_cell_a1 = ws1['A1']
    book1_cell_a1.fill = PatternFill(patternType='solid', fgColor=Color('cccccc'))
    book1_cell_b1 = ws1['B1']
    book1_cell_b1.fill = PatternFill(patternType='solid', fgColor=Color('cccccc'))
    book1_cell_c1 = ws1['C1']
    book1_cell_c1.fill = PatternFill(patternType='solid', fgColor=Color('cccccc'))
    book1_cell_d1 = ws1['D1']
    book1_cell_d1.fill = PatternFill(patternType='solid', fgColor=Color('cccccc'))
    book1_cell_e1 = ws1['E1']
    book1_cell_e1.fill = PatternFill(patternType='solid', fgColor=Color('cccccc'))
    book1_cell_f1 = ws1['F1']
    book1_cell_f1.fill = PatternFill(patternType='solid', fgColor=Color('cccccc'))
    book1_cell_g1 = ws1['G1']
    book1_cell_g1.fill = PatternFill(patternType='solid', fgColor=Color('cccccc'))
    book1.save('./ResultFile/response_table.xlsx')


# todo CWT 기간별 엑셀 저장 메소드
def cwt_term_excel_write(processing_df):
    yearly_n_niin_grouped = processing_df.groupby(['PLAN_YRMON', 'LOINO', 'N_NIIN']).mean()
    yearly_n_niin_grouped_in_10days = yearly_n_niin_grouped[yearly_n_niin_grouped['CWT'] <= 10]['CWT']
    yearly_n_niin_grouped_in_30days = \
        yearly_n_niin_grouped[(yearly_n_niin_grouped['CWT'] > 10) & (yearly_n_niin_grouped['CWT'] <= 30)]['CWT']
    yearly_n_niin_grouped_in_50days = \
        yearly_n_niin_grouped[(yearly_n_niin_grouped['CWT'] > 30) & (yearly_n_niin_grouped['CWT'] <= 50)]['CWT']
    yearly_n_niin_grouped_up_50days = yearly_n_niin_grouped[yearly_n_niin_grouped['CWT'] > 50]['CWT']
    p_yearly_n_niin_grouped = yearly_n_niin_grouped['CWT']
    yearly_n_niin_grouped_df = \
        pd.DataFrame(columns=[2013, 2014, 2015, 2016, 2017, 2018], index=['총 개수(개)', '10일 이내(%)', '30일 이내(%)', '50일 이내(%)', '50일 초과(%)'])
    yearly_n_niin_grouped_df.fillna(0, inplace=True)
    count_grouping_yearly(yearly_n_niin_grouped_df, p_yearly_n_niin_grouped, "총 개수(개)", 0)
    count_grouping_yearly(yearly_n_niin_grouped_df, yearly_n_niin_grouped_in_10days, "10일 이내(%)", 1)
    count_grouping_yearly(yearly_n_niin_grouped_df, yearly_n_niin_grouped_in_30days, "30일 이내(%)", 1)
    count_grouping_yearly(yearly_n_niin_grouped_df, yearly_n_niin_grouped_in_50days, "50일 이내(%)", 1)
    count_grouping_yearly(yearly_n_niin_grouped_df, yearly_n_niin_grouped_up_50days, "50일 초과(%)", 1)
    writer = pd.ExcelWriter('./ResultFile/cwt_by_period.xlsx', engine='openpyxl')
    if len(yearly_n_niin_grouped_in_10days) > 0:
        yearly_n_niin_grouped_in_10days.to_excel(writer, sheet_name='10일이내')
    if len(yearly_n_niin_grouped_in_30days) > 0:
        yearly_n_niin_grouped_in_30days.to_excel(writer, sheet_name='30일이내')
    if len(yearly_n_niin_grouped_in_50days) > 0:
        yearly_n_niin_grouped_in_50days.to_excel(writer, sheet_name='50일이내')
    if len(yearly_n_niin_grouped_up_50days) > 0:
        yearly_n_niin_grouped_up_50days.to_excel(writer, sheet_name='50일초과')
    if len(yearly_n_niin_grouped_df) > 0:
        yearly_n_niin_grouped_df.to_excel(writer, sheet_name='결과현황')
    writer.save()

    book = load_workbook('./ResultFile/cwt_by_period.xlsx')
    sheets = book.get_sheet_names()
    sheets.remove('결과현황')
    for sheetname in sheets:
        ws = book[sheetname]
        for cellname in ['A', 'B', 'C', 'D']:
            cell = ws[cellname + '1']
            cell.fill = PatternFill(patternType='solid', fgColor=Color('cccccc'))
        ws.column_dimensions['A'].width = 14.5
        ws.column_dimensions['B'].width = 16.5
        ws.column_dimensions['C'].width = 11.75
        ws.column_dimensions['D'].width = 12.13
    wsr = book['결과현황']
    for cellnamer in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
        cell = wsr[cellnamer + '1']
        cell.fill = PatternFill(patternType='solid', fgColor=Color('cccccc'))
    wsr.column_dimensions['A'].width = 12.75
    book.save('./ResultFile/cwt_by_period.xlsx')


try:
    # CWT Excel 파일
    target_file_path = read_ini('./dev.ini', 'target_file', 'file_path')
    sheetname = read_ini('./dev.ini', 'target_file', 'sheetname')

    # 공휴일 Excel 파일
    holiday_file_path = read_ini('./dev.ini', 'holiday_file', 'holiday_file_path')
    holiday_sheetname = read_ini('./dev.ini', 'holiday_file', 'holiday_sheetname')

    # excel에서 불러온 CWT 데이터셋 //엑셀파일불러오기 메소드 호출
    dataset = dataset_excelload(target_file_path, sheetname, 0)

    # dataset 재정렬한 딕셔너리//데이터셋 딕셔너리로 재정렬 메소드 호출
    processing_dataset, yearly_certificate_count, total_count = rearrangement_dataset(dataset)

    # excel에서 불러온 국가공휴일 데이터셋 //엑셀파일불러오기 메소드 호출
    holiday_dataset = dataset_excelload(holiday_file_path, holiday_sheetname, 1)

    # 연도별로 사전에 국가공휴일 정렬 //연도별 국가공휴일 정렬 메소드 호출
    holiday_processing_dataset = holiday_processing(holiday_dataset)

    # 총 정비기간 분석 메소드 호출
    processing_df = compare_temp(processing_dataset, holiday_processing_dataset)
    yearly_niin_count = yearly_niin_count(processing_df)

    # 요청사항 테이블 만드는 메소드 호출
    finally_df = finally_dataframe(processing_df, yearly_niin_count, yearly_certificate_count)

    # 엑셀 저장을 위한 디렉토리 생성
    mkdir_if_not('./ResultFile')

    # 요청사항 테이블 엑셀 저장 메소드 호출
    excel_write_finally_df(finally_df)

    # CWT 기간별 엑셀 저장 메소드 호출
    cwt_term_excel_write(processing_df)

    win32api.MessageBox(0, '프로그램이 정상적으로 완료 되었습니다.', '완료')
except:
    win32api.MessageBox(0, '프로그램 실행에 에러가 발생하였습니다.', '에러')
